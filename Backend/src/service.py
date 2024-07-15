"""Provides the timesheet template for the current month/year"""
import calendar
import os
import shutil
from collections import defaultdict
from datetime import datetime, timedelta
from io import BytesIO
from zipfile import ZipFile

import openpyxl
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, Side, numbers
from sqlalchemy.orm import Session

from src.db import db_reader as read
from src.db.model import models
from src.db.model.models import HolidayData, LeaveSheetData


def month_number_to_name(month_number):
    """Converts the month number to month name"""
    
    if 1 <= month_number <= 12:
        return calendar.month_name[month_number]
    else:
        return None


class TimeSheetTemplate:
    """Provides the basic structure for the time_sheet table."""

    def __init__(self, day_of_month: int, work_description: str = "", status: str = ""):
        self.day_of_month = day_of_month
        self.work_description = work_description
        self.status = status


def get_holidays(month: int, year: int, db: Session) -> dict:
    """Gets the holiday data from the table and returns the dictationary of it."""
    
    holidays = (
        db.query(HolidayData)
        .filter(
            HolidayData.holiday_date >= datetime(year, month, 1),
            HolidayData.holiday_date <= datetime(year, month, 28) + timedelta(days=3),
        )
        .all()
    )
    holiday_dict = {holiday.holiday_date.day: "Holiday" for holiday in holidays}
    return holiday_dict


def get_leaves(employee_id: int, month: int, year: int, db: Session) -> list:
    """Gets the leave data from the table and returns the dictationary of it."""
    
    leaves = (
        db.query(LeaveSheetData)
        .filter(
            LeaveSheetData.employee_id == employee_id,
            LeaveSheetData.transaction_status == "AVAILED",
            LeaveSheetData.leave_date >= datetime(year, month, 1),
            LeaveSheetData.leave_date <= datetime(year, month, 28) + timedelta(days=3),
        )
        .all()
    )
    return leaves


def create_timesheet_template(
    indxx_id: str, month: int, year: int, db: Session
) -> list:
    """Create a timesheet template for a given month and year
    Includes the Holiday, Leave and Weekends in the status column of the template."""

    days_in_month = (datetime(year, month % 12 + 1, 1) - timedelta(days=1)).day
    template = []

    empid = db.query(models.EmployeeData.employee_id).filter_by(indxx_id=indxx_id).first()
    if empid:
        employee_id = empid.employee_id
    else:
        raise ValueError("User with given Index ID does not exist.")
    
    holidays = get_holidays(month, year, db)
    leave_entries = get_leaves(employee_id, month, year, db)

    for day in range(1, days_in_month + 1):
        date = datetime(year, month, day)
        day_of_week = date.strftime("%A")

        if day_of_week in ["Saturday", "Sunday"]:
            status = day_of_week
        else:
            status = holidays.get(day, "")
        template.append(TimeSheetTemplate(day_of_month=day, status=status).__dict__)

    for leave in leave_entries:
        day_of_month = leave.leave_date.day

        for entry in template:
            if entry["day_of_month"] == day_of_month and entry["status"] == "":
                entry["status"] = "Leave"
    return template


thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def format_stoxx_timesheet(sh,i):
    """Formats the stoxx sheet according to template"""

    sh[f"A{i}"].alignment = Alignment(horizontal="center")
    sh[f"A{i}"].number_format = numbers.FORMAT_NUMBER
    sh[f"A{i}"].font = Font(size=9)
    sh[f"A{i}"].border = thin_border
    sh[f"A{i+1}"].border = thin_border
    sh[f"B{i}"].font = Font(size=10)
    sh[f"B{i}"].border = thin_border
    sh[f"B{i+1}"].border = thin_border
    sh[f"C{i}"].alignment = Alignment(horizontal="center")
    sh[f"C{i}"].border = thin_border
    sh[f"C{i+1}"].alignment = Alignment(horizontal="center")
    sh[f"C{i+1}"].font = Font(size=8)
    sh[f"C{i+1}"].border = thin_border
    sh[f"L{i}"].border = thin_border
    sh[f"L{i+1}"].border = thin_border
    sh[f"I{i}"].border = thin_border
    sh[f"I{i+1}"].border = thin_border
    sh[f"F{i}"].border = thin_border

    sh[f"F{i+1}"].border = thin_border
    sh[f"D{i}"].border = thin_border

    sh[f"D{i+1}"].border = thin_border
    sh[f"G{i}"].border = thin_border

    sh[f"G{i+1}"].border = thin_border
    sh[f"E{i}"].border = thin_border

    sh[f"E{i+1}"].border = thin_border
    sh[f"H{i}"].border = thin_border

    sh[f"H{i+1}"].border = thin_border
    sh[f"J{i}"].border = thin_border
    sh[f"J{i+1}"].border = thin_border
    sh[f"K{i}"].border = thin_border
    sh[f"K{i+1}"].border = thin_border
    sh[f"I{i}"].number_format = "0.00%"
    sh[f"I{i}"].alignment = Alignment(horizontal="center")
    sh[f"F{i}"].number_format = "#,##0.00"
    sh[f"F{i}"].alignment = Alignment(horizontal="center")
    sh[f"D{i}"].alignment = Alignment(horizontal="center")

    sh[f"G{i}"].number_format = "#,##0.00"
    sh[f"G{i}"].alignment = Alignment(horizontal="center")


    sh[f"E{i}"].number_format = "#,##0.00"
    sh[f"E{i}"].alignment = Alignment(horizontal="center")
    sh[f"H{i}"].alignment = Alignment(horizontal="center")
    sh[f"H{i}"].number_format = "0.00%"
    sh[f"K{i+1}"].number_format = (
                    '_("€"* #,##0.00_);_("€"* (#,##0.00);_("€"* "-"??_);_(@_)'
                )


def format_stoxx_sheet(ws,i):
    """Formats stoxx sheet"""
    ws[f'A{i}'].font = Font(name='Cambria',size=10)
    ws[f'B{i}'].font = Font(name='Cambria',size=10)
    ws[f'C{i}'].font = Font(name='Cambria',size=10)
    ws[f'D{i}'].font = Font(name='Cambria',size=10)
    ws[f'E{i}'].font = Font(name='Cambria',size=10)
    ws[f'F{i}'].font = Font(name='Cambria',size=10)
    ws[f'G{i}'].font = Font(name='Cambria',size=10)
    ws[f'H{i}'].font = Font(name='Cambria',size=10)
    ws[f'I{i}'].font = Font(name='Cambria',size=10)
    ws[f"A{i}"].border = thin_border
    ws[f"B{i}"].border = thin_border
    ws[f"C{i}"].border = thin_border
    ws[f"D{i}"].border = thin_border
    ws[f"E{i}"].border = thin_border
    ws[f"F{i}"].border = thin_border
    ws[f"G{i}"].border = thin_border
    ws[f"H{i}"].border = thin_border
    ws[f"I{i}"].border = thin_border
    ws[f"J{i}"].border = thin_border
    ws[f"K{i}"].border = thin_border
    ws[f"L{i}"].border = thin_border
    ws[f"M{i}"].border = thin_border
    ws[f"N{i}"].border = thin_border
    ws[f"O{i}"].border = thin_border
    ws[f"P{i}"].border = thin_border


def generate_stoxx_sheet(user,i,ws,leave_days):
    """Generate stoxx sheet"""
    
    proj_code=user.project_code.project_code
    ws['A2'] = user.team.team
    ws[f'A{i+4}'].value = i
    ws[f'B{i+4}'] = user.first_name
    ws[f'C{i+4}'] = user.last_name
    ws[f'D{i+4}'] = str(user.first_name + " " + user.last_name)
    ws[f'E{i+4}'] = user.hr_code
    ws[f'F{i+4}'] = user.team.team
    ws[f'G{i+4}'] = user.start_date
    ws[f'H{i+4}'] = user.level.level
    ws[f'I{i+4}'] = proj_code
    l=len(leave_days)
    if l == 0:
        ws[f'K{i+4}'] = '-'
    elif l == 1:
        days = str(tuple(leave_days))
        days = days.replace(',','')
        ws[f'K{i+4}'] = str(l) + " Day " + days
    else:
        days = str(tuple(leave_days))
        ws[f'K{i+4}'] = str(l) + " Days " + days
    format_stoxx_sheet(ws,i+4)


def generate_stoxx_timesheet(project_code_data, db: Session):
    """generates stoxx sheet for provided project_code, month and year"""
    
    project_code_list = project_code_data.project_code
    month = project_code_data.month
    year = project_code_data.year
    status_dict = defaultdict(list)
    upper_border = Border(top=Side(style="thin"))
    zip_buffer = BytesIO()
    with ZipFile(zip_buffer, "w") as zf:
        for project_code in project_code_list:
            proj_code_query = (
                db.query(models.EmployeeData)
                .join(models.ProjectCodeData)
                .filter(models.ProjectCodeData.project_code == project_code)
            )
            ofpx = "data/Stoxx_sheet_template.xlsx"
            nfpx = f"data/Stoxx_sheet_{project_code}.xlsx"
            if not os.path.exists(nfpx):
                shutil.copyfile(ofpx, nfpx)
            wbx = openpyxl.load_workbook(nfpx)
            wsx = wbx["stoxx_sheet"]
            wsx.title = project_code
            wsx = wbx[project_code]

            ofp = "data/Template_stoxx.xlsx"
            nfp = f"data/Template_stoxx_{project_code}.xlsx"
            if not os.path.exists(nfp):
                shutil.copyfile(ofp, nfp)
            wb = openpyxl.load_workbook(nfp)

            ws = wb["stoxx_sheet_template"]
            user_list = list(proj_code_query)
            for i in range(len(user_list)):
                wb.copy_worksheet(ws)

            ws = wb["stoxx_sheet_template Copy"]
            ws.title = "stoxx_sheet_template Copy0"

            wb.remove(wb["stoxx_sheet_template"])

            i = 9
            j = 1
            for user in proj_code_query:

                sh = wb["Summary"]
                ws = wb[f"stoxx_sheet_template Copy{j-1}"]

                sh[f"A{i}"].value = j
                ws.title = str(user.hr_code)
                sh[f"B{i}"] = str(user.hr_code)
                ws["C6"] = str(user.first_name + " " + user.last_name)
                sh[f"C{i}"] = ws["C6"].value
                sh[f"C{i+1}"] = "Travel"
                ws["J2"] = str(str(month_number_to_name(month)) + "," + str(year))
                sh["E3"] = str(str(month_number_to_name(month)) + "," + str(year))
                sh["E8"] = str(str(month_number_to_name(month)) + "," + str(year))
                ws["C8"] = user.project_number.project_number
                ws["I4"] = user.manager.manager
                ws["I6"] = user.department.department
                ws["I10"] = user.project_name.project_name
                sh["E4"] = user.project_name.project_name
                timesheet_data = read.get_timesheet_by_indxx_id_and_date(
                    db, str(user.indxx_id), month, year
                )
                ws.sheet_view.showGridLines = False
                format_stoxx_timesheet(sh,i)

                if j == len(user_list):
                    i = i + 4
                    sh[f"C{i}"].border = upper_border
                    sh[f"D{i}"].border = upper_border
                    sh[f"E{i}"].border = upper_border

                    sh[f"C{i}"] = "Name"
                    sh[f"C{i}"].alignment = Alignment(horizontal="left")
                    sh[f"E{i}"] = "Date"
                    sh[f"E{i}"].alignment = Alignment(horizontal="left")
                    sh[f"J{i}"] = "Name"
                    sh[f"J{i}"].alignment = Alignment(horizontal="left")
                    sh[f"J{i}"].font = Font(size=10)

                    sh[f"K{i}"] = "Date"
                    sh[f"K{i}"].alignment = Alignment(horizontal="left")
                    sh[f"K{i}"].font = Font(size=10)

                    i = i + 2

                    sh[f"C{i}"] = "Signature indxx line manager"
                    sh[f"C{i}"].alignment = Alignment(horizontal="left")
                    sh[f"J{i}"] = "Signature STOXX line manager"
                    sh[f"J{i}"].alignment = Alignment(horizontal="left")
                    sh[f"J{i}"].font = Font(size=10)
                    i = i - 6

                if not timesheet_data:
                    status_dict[project_code].append(str(user.first_name + " " + user.last_name))
                    wb.save(
                        f"data/Template_stoxx_{user.project_code.project_code}.xlsx"
                    )
                    i = i + 2
                    leave_days = []
                    generate_stoxx_sheet(user,j,wsx,leave_days)
                    wb.save(nfp)
                    wbx.save(nfpx)
                    j = j + 1
                    continue
                df = pd.DataFrame(timesheet_data)
                ws["L4"] = str(df.loc[0]["work_description"])
                sh[f"L{i}"] = ws["L4"].value
                sh[f"I{i}"].value = 1
                df = df.reset_index(drop=True)
                df = df.drop(0)

                df["IN"] = ""
                df.loc[df["status"].isnull() | (df["status"] == ""), "IN"] = "10:00"
                df["OUT"] = ""
                df.loc[df["status"].isnull() | (df["status"] == ""), "OUT"] = "19:00"
                df["total_hrs"] = ""
                df.loc[df["status"].isnull() | (df["status"] == ""), "total_hrs"] = (
                    8.00
                )
                df["break_hrs"] = ""
                df.loc[df["status"].isnull() | (df["status"] == ""), "break_hrs"] = (
                    "0:30"
                )

                df["status"].replace({"Saturday": "", "Sunday": ""}, inplace=True)
                df["working_day"] = 0
                df.loc[df["IN"] != "", "working_day"] = 1
                leave_days = df[df['status'] == 'Leave']['day_of_month'].to_list()
                leaves_count = df["status"].value_counts().get("Leave", 0)

                startrow = 14
                startcols = {
                    "work_description": 12,
                    "IN": 1,
                    "OUT": 2,
                    "status": 8,
                    "total_hrs": 4,
                    "break_hrs": 3,
                    "working_day": 6,
                }

                for idx, row in df.iterrows():
                    for col, startcol in startcols.items():
                        value = row[col]
                        cell_row = startrow + int(idx)
                        cell_col = startcol + 1
                        ws.cell(row=cell_row, column=cell_col, value=value)
                  
                count = df["working_day"].sum()

                ws["G48"] = count
                ws["E48"] = 8 * count
                ws["E50"] = count
                ws["G50"] = count
                ws["D57"] = count
                ws["D58"] = -count

                sh[f"F{i}"] = ws["E50"].value

                sh[f"D{i}"] = ws["E48"].value

                sh[f"G{i}"] = ws["G50"].value

                sh[f"E{i}"] = count + leaves_count

                sh[f"H{i}"] = count / (count + leaves_count)
                if sh[f"J{i+1}"].value is not None and sh[f"I{i+1}"].value is not None:
                    sh[f"K{i+1}]"].value = int(sh[f"J{i+1}"].value) * int(
                        sh[f"I{i+1}"].value
                    )
                else:
                    sh[f"K{i+1}"].value = 0

                i = i + 2
                generate_stoxx_sheet(user,j,wsx,leave_days)
                wbx.save(nfpx)
                j = j + 1
                wb.save(nfp)
            
            
            zf.write(nfp, os.path.join(project_code,os.path.basename(nfp)))
            zf.write(nfpx, os.path.join(project_code,os.path.basename(nfpx)))
            os.remove(nfp)
            os.remove(nfpx)
    zip_buffer.seek(0)
    if not status_dict:
        status_list = []
        status_list.append({"details": "Generated successfully"})
    else:
        status_list = [{"details": f"{proj_code}: {', '.join(indxx_ids)} have not filled the timesheet"}
                  for proj_code, indxx_ids in status_dict.items()]
    return zip_buffer.getvalue(), status_list